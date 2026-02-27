import requests
import pandas as pd
import urllib.parse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from tqdm import tqdm
from colorama import init, Fore, Style
import time
import os
import datetime

init(autoreset=True)

# === НАСТРОЙКИ (Cookies) ===
HEADERS = {
    'Accept': '*/*',
    'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
    'Connection': 'keep-alive',
    'Origin': 'https://www.wildberries.ru',
    'Referer': 'https://www.wildberries.ru/',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:147.0) Gecko/20100101 Firefox/147.0',
    'x-queryid': 'YOUR_QUERY_ID_HERE', 
    'x-requested-with': 'XMLHttpRequest',
    'Cookie': 'YOUR_COOKIE_HERE',
    'Sec-Fetch-Dest': 'empty',
    'Sec-Fetch-Site': 'same-origin',
    'Priority': 'u=4',
}

def get_products_by_query(query, pages=1):
    all_items = []
    encoded_query = urllib.parse.quote(query)
    
    print(f"\n{Fore.CYAN}🚀 Запуск сканирования: {Style.BRIGHT}'{query}'{Style.RESET_ALL}")
    
    for page in tqdm(range(1, pages + 1), desc="Загрузка данных", unit="стр", colour="green"):
        url = f"https://www.wildberries.ru/__internal/u-search/exactmatch/ru/common/v18/search?ab_testing=false&appType=1&curr=rub&dest=-1257786&hide_dtype=9&hide_vflags=4294967296&lang=ru&page={page}&query={encoded_query}&resultset=catalog&sort=popular&spp=30&suppressSpellcheck=false"
        
        try:
            time.sleep(0.3)
            response = requests.get(url, headers=HEADERS, timeout=15)
            response.raise_for_status()
            data = response.json()
            
            items = data.get('products', [])
            if not items:
                items = data.get('data', {}).get('products', [])
            
            if not items:
                break
                
            for item in items:
                price = 0
                try:
                    price = item['sizes'][0]['price']['product'] / 100
                except:
                    price = item.get('salePriceU', 0) / 100

                rating = item.get('reviewRating', item.get('rating', 0))
                feedbacks = item.get('feedbacks', 0)
                
                # Считаем LTV (Lifetime Value) - общая выручка карточки за всё время
                revenue_estimate = price * feedbacks
                
                all_items.append({
                    'ID': item.get('id'),
                    'Бренд': item.get('brand', ''),
                    'Название': item.get('name', ''),
                    'Цена': price,
                    'Рейтинг': rating,
                    'Отзывы': feedbacks,
                    'Выручка (LTV)*': revenue_estimate, 
                    'Ссылка': f"https://www.wildberries.ru/catalog/{item.get('id')}/detail.aspx"
                })
                
        except Exception:
            continue
            
    return all_items

def create_analytics_sheet(wb, df):
    ws = wb.create_sheet("📊 АНАЛИТИКА НИШИ")
    
    # 1. Расчет метрик
    avg_price = df['Цена'].mean()
    total_revenue_ltv = df['Выручка (LTV)*'].sum()
    
    # Топ-3 бренда
    try:
        top_brands_list = df['Бренд'].value_counts().head(3).index.tolist()
        top_brands_str = ", ".join(top_brands_list)
    except:
        top_brands_str = "Нет данных"

    # Дата отчета
    scan_date = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
    
    data = [
        ("ПАРАМЕТР", "ЗНАЧЕНИЕ"),
        ("Дата сканирования", scan_date),
        ("Всего товаров в выборке", len(df)),
        ("Средняя цена товара", f"{avg_price:.0f} ₽"),
        ("ТОП-3 Бренда (доля рынка)", top_brands_str),
        ("Оборот ниши (накопительный LTV*)", f"{total_revenue_ltv:,.0f} ₽"),
        ("Самый дорогой товар", f"{df['Цена'].max():.0f} ₽"),
        ("Самый дешевый товар", f"{df['Цена'].min():.0f} ₽"),
        ("", "* LTV (Lifetime Value) — оценка выручки за всё время существования карточек (Цена × Отзывы).")
    ]
    
    for row in data:
        ws.append(row)
        
    # Форматирование
    header_fill = PatternFill(start_color="B00000", end_color="B00000", fill_type="solid") # Темно-красный
    white_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Красим шапку
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font
    
    # Ширина и сноска
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 40
    
    # Сноска серым цветом
    ws['A9'].font = Font(italic=True, color="808080", size=9)

def save_smart_excel(data, filename):
    if not data: return
    
    df = pd.DataFrame(data)
    # Сортируем по популярности (Выручка LTV)
    df = df.sort_values(by='Выручка (LTV)*', ascending=False)
    
    # Сохраняем и форматируем
    try:
        df.to_excel(filename, index=False)
        wb = load_workbook(filename)
        ws = wb.active
        ws.title = "Товары"
        
        create_analytics_sheet(wb, df)
        
        # Стили главного листа
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        widths = {'A': 15, 'B': 20, 'C': 50, 'D': 12, 'E': 10, 'F': 10, 'G': 20, 'H': 15}
        for col_name, width in widths.items():
            ws.column_dimensions[col_name].width = width

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical='center')
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                # Денежный формат
                if cell.col_idx in [4, 7]: 
                    cell.number_format = '#,##0 "₽"'
                
                # Ссылка "Перейти"
                if cell.col_idx == 8 and cell.row > 1:
                    cell.hyperlink = cell.value
                    cell.value = "Перейти >>"
                    cell.font = Font(color="0563C1", underline="single")

        ws.auto_filter.ref = ws.dimensions
        wb.save(filename)
        print(f"\n{Fore.GREEN}✅ Файл сохранен: {Style.BRIGHT}{filename}")
        print(f"{Fore.YELLOW}📊 Лист аналитики сформирован успешно.")
        
    except PermissionError:
        print(f"\n{Fore.RED}[!] Ошибка: Закройте Excel файл и попробуйте снова!")

def main():
    os.system('cls' if os.name == 'nt' else 'clear')
    print(f"{Fore.MAGENTA}{Style.BRIGHT}=== WB ANALYTICS TOOL v6.1 (Portfolio Ed.) ==={Style.RESET_ALL}")
    
    user_query = input("Введите поисковый запрос: ").strip()
    if not user_query: return

    try:
        pages = int(input("Глубина поиска (страниц, Enter=3): ") or 3)
    except:
        pages = 3
        
    filename = f"WB_{user_query.replace(' ', '_')}.xlsx"
    products = get_products_by_query(user_query, pages)
    
    if products:
        save_smart_excel(products, filename)
    else:
        print(f"\n{Fore.RED}[!] Данные не получены. Проверьте Cookies.")
        
    input("\nНажмите Enter, чтобы закрыть...")

if __name__ == '__main__':
    main()